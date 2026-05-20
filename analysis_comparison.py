"""
Comparison engine for ETABS Analysis Output Tables (Element Forces).

Parses Element Forces - Columns/Beams/Braces sheets from two Excel exports
and produces a member-level max-demand-envelope comparison.

No design D/C ratios — pure force envelope. One threshold band for WARN,
one for FAIL; no gravity/lateral distinction since combo names are not used.
"""
import hashlib
import io

import pandas as pd
from openpyxl import load_workbook

from comparison import _stream_etabs_rows, _pct_change, _fmt
from storage_cache import get_cached, set_cached

ANALYSIS_FORCES_SHEETS = {
    'Columns': 'Element Forces - Columns',
    'Beams':   'Element Forces - Beams',
    'Braces':  'Element Forces - Braces',
}

ANALYSIS_LABEL_COL = {
    'Columns': 'Column',
    'Beams':   'Beam',
    'Braces':  'Brace',
}

# Singular display names matching design-mode convention
MEMBER_TYPE_LABEL = {
    'Columns': 'Column',
    'Beams':   'Beam',
    'Braces':  'Brace',
}

ALL_FORCE_COMPONENTS = ['P', 'V2', 'V3', 'T', 'M2', 'M3']

# Only these forces are checked per member type — others stored as N/A
GOVERNING_FORCES = {
    'Columns': ['P', 'M3', 'M2'],
    'Beams':   ['M3', 'V2'],
    'Braces':  ['P'],
}

# Caching handled by storage_cache. Prefixes: 'etabs_aparsev3' / 'etabs_acmpv3'.
# v2: adds OutputCase tracking (governing combo) to parsed DataFrames.


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def _parse_analysis_forces_from_wb(wb, member_type: str) -> pd.DataFrame:
    """
    Stream one Element Forces worksheet and return one row per member with
    max-absolute-value envelope across all combos, stations, and step types.
    """
    sheet_name     = ANALYSIS_FORCES_SHEETS[member_type]
    label_col_name = ANALYSIS_LABEL_COL[member_type]
    combo_cols = [f'{c}_Combo' for c in ALL_FORCE_COMPONENTS]
    empty = pd.DataFrame(columns=['Story', 'Label'] + ALL_FORCE_COMPONENTS + combo_cols)

    if sheet_name not in wb.sheetnames:
        return empty

    try:
        col_map, rows_iter = _stream_etabs_rows(wb[sheet_name])
    except StopIteration:
        return empty

    story_i      = col_map.get('Story')
    label_i      = col_map.get(label_col_name)
    # ETABS column name varies by version: 'Output Case', 'OutputCase', 'Load Case/Combo'
    _oc_candidates = ['Output Case', 'OutputCase', 'Load Case/Combo']
    output_case_i = next((col_map[k] for k in _oc_candidates if k in col_map), None)
    if story_i is None or label_i is None:
        return empty

    force_col_indices = {c: col_map.get(c) for c in ALL_FORCE_COMPONENTS}
    accum: dict = {}

    for row in rows_iter:
        vals = [cell.value for cell in row]
        n     = len(vals)
        story = vals[story_i] if story_i < n else None
        label = vals[label_i] if label_i < n else None
        if story is None or label is None:
            continue
        output_case = (
            str(vals[output_case_i])
            if output_case_i is not None and output_case_i < n and vals[output_case_i] is not None
            else ''
        )
        key = (story, label)
        if key not in accum:
            accum[key] = {c: 0.0 for c in ALL_FORCE_COMPONENTS}
            accum[key].update({f'{c}_Combo': '' for c in ALL_FORCE_COMPONENTS})
        for c, ci in force_col_indices.items():
            if ci is not None and ci < n and vals[ci] is not None:
                try:
                    v = float(vals[ci])
                    if abs(v) > abs(accum[key][c]):
                        accum[key][c] = v
                        accum[key][f'{c}_Combo'] = output_case
                except (TypeError, ValueError):
                    pass

    if not accum:
        return empty
    return pd.DataFrame(
        [{'Story': s, 'Label': l, **forces} for (s, l), forces in accum.items()]
    )


def _get_parsed_analysis_file(file_obj) -> tuple:
    """Return (hash, parsed_dict), computing and caching on first call."""
    file_bytes = file_obj.getvalue_binary()
    h = hashlib.md5(file_bytes).hexdigest()

    cached = get_cached('etabs_aparsev3', h)
    if cached is not None:
        return h, cached

    fh = io.BytesIO(file_bytes)
    del file_bytes
    try:
        wb = load_workbook(fh, read_only=True, data_only=True)
        try:
            result = {
                'forces': {
                    mt: _parse_analysis_forces_from_wb(wb, mt)
                    for mt in ('Columns', 'Beams', 'Braces')
                },
                'sheet_names': list(wb.sheetnames),
            }
        finally:
            wb.close()
    except Exception:
        result = {
            'forces': {
                mt: pd.DataFrame(columns=['Story', 'Label'] + ALL_FORCE_COMPONENTS)
                for mt in ('Columns', 'Beams', 'Braces')
            },
            'sheet_names': [],
        }

    set_cached('etabs_aparsev3', h, result)
    return h, result


# ---------------------------------------------------------------------------
# Comparison
# ---------------------------------------------------------------------------

def _run_analysis_internal(
    parsed_exist: dict,
    parsed_new: dict,
    member_type_filter: str,
    warn_threshold: float,
    fail_threshold: float,
) -> list:
    types_to_process = (
        ['Columns', 'Beams', 'Braces']
        if member_type_filter == 'All'
        else [member_type_filter]
    )

    all_results = []

    for mtype in types_to_process:
        gov_forces   = GOVERNING_FORCES[mtype]
        display_type = MEMBER_TYPE_LABEL[mtype]
        forces_exist = parsed_exist['forces'][mtype]
        forces_new   = parsed_new['forces'][mtype]

        def _member_set(df):
            if df.empty or 'Story' not in df.columns:
                return set()
            return set(zip(df['Story'], df['Label']))

        exist_members = _member_set(forces_exist)
        new_members   = _member_set(forces_new)
        all_members   = exist_members | new_members

        def _build_lookup(df):
            out = {}
            if df.empty:
                return out
            for _, r in df.iterrows():
                out[(r['Story'], r['Label'])] = r
            return out

        fe_lookup = _build_lookup(forces_exist)
        fn_lookup = _build_lookup(forces_new)

        def _gov_combo(series, forces: list) -> str:
            """Return the OutputCase that drove the max absolute governing force."""
            if series is None:
                return ''
            best_val, best_combo = 0.0, ''
            for c in forces:
                try:
                    v = abs(float(series[c]))
                    if v > best_val:
                        best_val = v
                        best_combo = str(series.get(f'{c}_Combo', '') or '')
                except (TypeError, ValueError, KeyError):
                    pass
            return best_combo or 'N/A'

        for story, label in sorted(all_members, key=lambda x: (str(x[0]), str(x[1]))):
            in_exist = (story, label) in exist_members
            in_new   = (story, label) in new_members
            ef = fe_lookup.get((story, label))
            nf = fn_lookup.get((story, label))

            row = {
                'Story':      story,
                'Label':      label,
                'MemberType': display_type,
            }

            # Pre-populate all force slots as N/A; governing forces overwrite below
            for c in ALL_FORCE_COMPONENTS:
                row[f'{c}_Exist'] = 'N/A'
                row[f'{c}_New']   = 'N/A'
                row[f'{c}_Pct']   = 'N/A'

            if not in_new:
                for c in gov_forces:
                    row[f'{c}_Exist'] = _fmt(float(ef[c]), 2) if ef is not None else 'N/A'
                row.update({
                    'GovCombo_Exist': _gov_combo(ef, gov_forces),
                    'GovCombo_New':   'N/A',
                    'WorstPct': 'N/A', 'FailReason': '', 'Pass': 'REMOVED',
                })
                all_results.append(row)
                continue

            if not in_exist:
                for c in gov_forces:
                    row[f'{c}_New'] = _fmt(float(nf[c]), 2) if nf is not None else 'N/A'
                row.update({
                    'GovCombo_Exist': 'N/A',
                    'GovCombo_New':   _gov_combo(nf, gov_forces),
                    'WorstPct': 'N/A', 'FailReason': '', 'Pass': 'ADDED',
                })
                all_results.append(row)
                continue

            # Matched — compute % change for governing forces only
            worst_pct   = None
            worst_force = None

            for c in gov_forces:
                e_val = _fmt(float(ef[c]), 2) if ef is not None else 'No Data'
                n_val = _fmt(float(nf[c]), 2) if nf is not None else 'No Data'
                row[f'{c}_Exist'] = e_val
                row[f'{c}_New']   = n_val

                if isinstance(e_val, (int, float)) and isinstance(n_val, (int, float)):
                    pct, _ = _pct_change(float(e_val), float(n_val))
                    row[f'{c}_Pct'] = pct
                    numeric = pct if isinstance(pct, float) else (999.0 if pct == 'INF' else None)
                    if numeric is not None and (worst_pct is None or numeric > worst_pct):
                        worst_pct   = numeric
                        worst_force = c
                else:
                    row[f'{c}_Pct'] = 'N/A'

            row['GovCombo_Exist'] = _gov_combo(ef, gov_forces)
            row['GovCombo_New']   = _gov_combo(nf, gov_forces)
            row['WorstPct'] = round(worst_pct, 1) if worst_pct is not None else 'N/A'

            if worst_pct is not None and worst_pct > fail_threshold:
                pct_val = row[f'{worst_force}_Pct']
                pct_str = f'+{pct_val:.1f}%' if isinstance(pct_val, float) else 'INF'
                row['FailReason'] = f'{worst_force} {pct_str} > {fail_threshold:.0f}%'
                row['Pass'] = 'FLAG'
            elif worst_pct is not None and worst_pct > warn_threshold:
                pct_val = row[f'{worst_force}_Pct']
                pct_str = f'+{pct_val:.1f}%' if isinstance(pct_val, float) else str(pct_val)
                row['FailReason'] = f'{worst_force} {pct_str} > {warn_threshold:.0f}%'
                row['Pass'] = 'WARN'
            else:
                row['FailReason'] = ''
                row['Pass'] = 'PASS'

            all_results.append(row)

    return all_results


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def run_analysis_comparison(
    existing_file,
    modified_file,
    member_type_filter: str,
    warn_threshold: float,
    fail_threshold: float,
    show_failures_only: bool,
) -> list:
    """
    Compare ETABS element force envelopes between two models.

    Returns one row per member. Keys per row: Story, Label, MemberType,
    P/V2/M2/M3_Exist/New/Pct (governing forces only; others N/A),
    WorstPct, FailReason, Pass.
    """
    exist_hash, parsed_exist = _get_parsed_analysis_file(existing_file)
    new_hash,   parsed_new   = _get_parsed_analysis_file(modified_file)
    cache_key = (exist_hash, new_hash, member_type_filter, warn_threshold, fail_threshold)

    all_results = get_cached('etabs_acmpv3', cache_key)
    if all_results is None:
        all_results = _run_analysis_internal(
            parsed_exist, parsed_new,
            member_type_filter, warn_threshold, fail_threshold,
        )
        set_cached('etabs_acmpv3', cache_key, all_results)

    if show_failures_only:
        return [r for r in all_results if r.get('Pass') in ('FLAG', 'WARN')]
    return list(all_results)

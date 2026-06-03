"""
Comparison engine for ETABS Design Output Tables.

Parses Design Forces and Steel Frame Design Summary sheets from two Excel
exports and produces a member-level force and D/C ratio comparison table.

Performance: parsed DataFrames and comparison results are cached in
module-level dicts keyed on MD5 hash of file content, so re-running with
different thresholds or display filters is near-instant after the first parse.
"""
import hashlib
import io
import re
from collections import defaultdict

import pandas as pd
from openpyxl import load_workbook

from storage_cache import get_cached, set_cached

FORCE_COMPONENTS = ['P', 'V2', 'V3', 'T', 'M2', 'M3']

GOVERNING_FORCES = {
    'Columns': ['P', 'M3', 'M2'],
    'Beams':   ['M3', 'V2'],
    'Braces':  ['P'],
}

DESIGN_FORCES_SHEETS = {
    'Columns': 'Design Forces - Columns',
    'Beams':   'Design Forces - Beams',
    'Braces':  'Design Forces - Braces',
}

# Member-specific label column name in each design forces sheet
LABEL_COL = {
    'Columns': 'Column',
    'Beams':   'Beam',
    'Braces':  'Brace',
}

SUMMARY_SHEET_PREFIX = 'Stl Frm Sum'
COMPOSITE_BEAM_SUMMARY_PREFIX = 'Comp Bm Sum'

# Design type string as it appears in the Design Summary sheet
DESIGN_TYPE = {
    'Columns': 'Column',
    'Beams':   'Beam',
    'Braces':  'Brace',
}

# Lateral load token: E or W followed by one or more letters (e.g. EQ, EQB, WA, WG).
# Lookbehind excludes letters only so "0.5WG" and "1EQB" (digit-adjacent) still match,
# while "Dead", "Lr", "SB" etc. don't start with E/W and are unaffected.
LATERAL_RE = re.compile(r'(?<![A-Za-z])([EW][A-Za-z]+)(?![A-Za-z])')


# Caching is handled by storage_cache (two-level: in-process + Viktor Storage).
# Prefixes: 'etabs_parsev3' for parsed workbook data, 'etabs_cmpv4' for comparison results.


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def _stream_etabs_rows(ws):
    """
    Return (col_map, rows_iter) for an ETABS-formatted worksheet.

    ETABS layout: row 0 = title, row 1 = headers, row 2 = units, row 3+ = data.
    col_map maps column header name → 0-based column index.
    Raises StopIteration if the sheet has fewer than 3 rows.
    """
    rows_iter = ws.rows
    next(rows_iter)                      # row 0: table title
    header_cells = next(rows_iter)       # row 1: column headers
    next(rows_iter)                      # row 2: units — skip
    col_map = {
        cell.value: i
        for i, cell in enumerate(header_cells)
        if cell.value is not None
    }
    return col_map, rows_iter


# ---------------------------------------------------------------------------
# Workbook-level parsers (take an already-open read_only workbook)
# ---------------------------------------------------------------------------

def _parse_forces_from_wb(wb, member_type: str) -> pd.DataFrame:
    """
    Stream one Design Forces worksheet from an open workbook.

    Online max-abs aggregation: stores one float per (Story, Label, component).
    Also tracks which OutputCase drove the max for each component ({c}_Combo).
    Peak memory is O(unique members), not O(rows), regardless of file size.
    """
    sheet_name    = DESIGN_FORCES_SHEETS[member_type]
    label_col_name = LABEL_COL[member_type]
    combo_cols = [f'{c}_Combo' for c in FORCE_COMPONENTS]
    empty = pd.DataFrame(columns=['Story', 'Label'] + FORCE_COMPONENTS + combo_cols)

    if sheet_name not in wb.sheetnames:
        return empty

    try:
        col_map, rows_iter = _stream_etabs_rows(wb[sheet_name])
    except StopIteration:
        return empty

    story_i = col_map.get('Story')
    label_i = col_map.get(label_col_name)
    if story_i is None or label_i is None:
        return empty

    # Design Forces sheets use 'Combo' or 'Load Combination'; Element Forces use 'Output Case'
    _oc_candidates = ['Combo', 'Load Combination', 'Output Case', 'OutputCase', 'Load Case/Combo']
    output_case_i = next((col_map[k] for k in _oc_candidates if k in col_map), None)

    force_col_indices = {c: col_map.get(c) for c in FORCE_COMPONENTS}
    accum: dict = {}

    for row in rows_iter:
        vals = [cell.value for cell in row]
        n    = len(vals)
        story = vals[story_i] if story_i < n else None
        label = vals[label_i] if label_i < n else None
        if story is None or label is None:
            continue
        output_case = (
            str(vals[output_case_i])
            if output_case_i is not None and output_case_i < n and vals[output_case_i] is not None
            else ''
        )
        key = (story, label)
        if key not in accum:
            accum[key] = {c: 0.0 for c in FORCE_COMPONENTS}
            accum[key].update({f'{c}_Combo': '' for c in FORCE_COMPONENTS})
        for c, ci in force_col_indices.items():
            if ci is not None and ci < n and vals[ci] is not None:
                try:
                    v = float(vals[ci])
                    if abs(v) > abs(accum[key][c]):
                        accum[key][c] = v
                        accum[key][f'{c}_Combo'] = output_case
                except (TypeError, ValueError):
                    pass

    if not accum:
        return empty
    return pd.DataFrame(
        [{'Story': s, 'Label': l, **forces} for (s, l), forces in accum.items()]
    )


_SUMMARY_EMPTY = pd.DataFrame(columns=[
    'Story', 'Label', 'MemberType', 'DesignSection',
    'PMMCombo', 'PMMRatio', 'VMajCombo', 'VMajRatio',
])

_SUMMARY_NAME_REMAP = {
    'Design Type':    'MemberType',
    'Design Section': 'DesignSection',
    'PMM Combo':      'PMMCombo',
    'PMM Ratio':      'PMMRatio',
    'V Major Combo':  'VMajCombo',
    'V Major Ratio':  'VMajRatio',
}


def _parse_one_summary_sheet(ws) -> pd.DataFrame:
    """Parse a single design-summary worksheet (steel or composite) into a DataFrame."""
    try:
        col_map, rows_iter = _stream_etabs_rows(ws)
    except StopIteration:
        return _SUMMARY_EMPTY.copy()

    idx = {_SUMMARY_NAME_REMAP.get(k, k): i for k, i in col_map.items()}
    story_i = idx.get('Story')
    label_i = idx.get('Label')
    if story_i is None or label_i is None:
        return _SUMMARY_EMPTY.copy()

    extra_fields = ('MemberType', 'DesignSection', 'PMMCombo', 'PMMRatio', 'VMajCombo', 'VMajRatio')
    rows_out = []
    for row in rows_iter:
        vals = [cell.value for cell in row]
        n    = len(vals)
        story = vals[story_i] if story_i < n else None
        label = vals[label_i] if label_i < n else None
        if story is None or label is None:
            continue
        row_dict = {'Story': story, 'Label': label}
        for field in extra_fields:
            fi = idx.get(field)
            row_dict[field] = vals[fi] if fi is not None and fi < n else None
        rows_out.append(row_dict)

    if not rows_out:
        return _SUMMARY_EMPTY.copy()

    df = pd.DataFrame(rows_out)
    df = df[df['Label'].notna()].copy()
    for col in ('PMMCombo', 'VMajCombo'):
        if col in df.columns:
            df[col] = (
                df[col].astype(str)
                .str.replace(r'\s*\([CT]\)\s*$', '', regex=True)
                .str.strip()
            )
    for col in ('PMMRatio', 'VMajRatio'):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


def _parse_summary_from_wb(wb) -> pd.DataFrame:
    """
    Stream design summary worksheet(s) from an open workbook.

    Reads Steel Frame Design Summary (prefix 'Stl Frm Sum') and, when present,
    Composite Beam Design Summary (prefix 'Comp Bm Sum').  Composite beam rows
    are normalised to MemberType='Beam' so the existing comparison logic handles
    them without modification.  Both sheets share the same column schema.
    """
    frames = []

    steel_sheet = next(
        (s for s in wb.sheetnames if s.startswith(SUMMARY_SHEET_PREFIX)), None
    )
    if steel_sheet is not None:
        frames.append(_parse_one_summary_sheet(wb[steel_sheet]))

    comp_sheet = next(
        (s for s in wb.sheetnames if s.startswith(COMPOSITE_BEAM_SUMMARY_PREFIX)), None
    )
    if comp_sheet is not None:
        comp_df = _parse_one_summary_sheet(wb[comp_sheet])
        if not comp_df.empty and 'MemberType' in comp_df.columns:
            comp_df['MemberType'] = 'Beam'
        frames.append(comp_df)

    if not frames:
        return _SUMMARY_EMPTY.copy()

    combined = pd.concat(frames, ignore_index=True)
    return combined if not combined.empty else _SUMMARY_EMPTY.copy()


# ---------------------------------------------------------------------------
# Public parser wrappers (open their own workbook — used for standalone testing)
# ---------------------------------------------------------------------------

def parse_design_forces(file_source, member_type: str) -> pd.DataFrame:
    """Parse one Design Forces sheet. _get_parsed_file uses _parse_forces_from_wb instead."""
    try:
        if isinstance(file_source, (bytes, bytearray)):
            fh = io.BytesIO(file_source)
        else:
            with file_source.open_binary() as raw:
                fh = io.BytesIO(raw.read())
        wb = load_workbook(fh, read_only=True, data_only=True)
        try:
            return _parse_forces_from_wb(wb, member_type)
        finally:
            wb.close()
    except Exception:
        return pd.DataFrame(columns=['Story', 'Label'] + FORCE_COMPONENTS)


def parse_design_summary(file_source) -> pd.DataFrame:
    """Parse the Design Summary sheet. _get_parsed_file uses _parse_summary_from_wb instead."""
    try:
        if isinstance(file_source, (bytes, bytearray)):
            fh = io.BytesIO(file_source)
        else:
            with file_source.open_binary() as raw:
                fh = io.BytesIO(raw.read())
        wb = load_workbook(fh, read_only=True, data_only=True)
        try:
            return _parse_summary_from_wb(wb)
        finally:
            wb.close()
    except Exception:
        return pd.DataFrame()


# ---------------------------------------------------------------------------
# Cache helpers
# ---------------------------------------------------------------------------

def _get_parsed_file(file_obj) -> tuple:
    """
    Return (file_hash, parsed_dict) for a file, computing and caching if absent.

    Opens the workbook exactly once and reads all sheets in a single pass,
    including coord index extraction for By Coordinates matching.

    parsed_dict = {
        'summary': df,
        'forces': {'Columns': df, 'Beams': df, 'Braces': df},
        'sheet_names': [...],
        'coord_index': {label: geom_key, ...},   # empty dict if sheets absent
    }
    """
    from coord_matching import parse_coord_index

    file_bytes = file_obj.getvalue_binary()
    h = hashlib.md5(file_bytes).hexdigest()

    cached = get_cached('etabs_parsev4', h)
    if cached is not None:
        return h, cached

    fh = io.BytesIO(file_bytes)
    del file_bytes
    try:
        wb = load_workbook(fh, read_only=True, data_only=True)
        try:
            result = {
                'summary': _parse_summary_from_wb(wb),
                'forces': {
                    mt: _parse_forces_from_wb(wb, mt)
                    for mt in ('Columns', 'Beams', 'Braces')
                },
                'sheet_names': list(wb.sheetnames),
                'coord_index': parse_coord_index(wb),
            }
        finally:
            wb.close()
    except Exception:
        _combo_cols = [f'{c}_Combo' for c in FORCE_COMPONENTS]
        result = {
            'summary': pd.DataFrame(),
            'forces': {
                mt: pd.DataFrame(columns=['Story', 'Label'] + FORCE_COMPONENTS + _combo_cols)
                for mt in ('Columns', 'Beams', 'Braces')
            },
            'sheet_names': [],
            'coord_index': {},
        }

    set_cached('etabs_parsev4', h, result)
    return h, result


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

def classify_combo(combo_name: str) -> str:
    """
    Return 'lateral' if the combo name contains a seismic (E...) or wind (W...)
    load token; otherwise 'gravity'.

    ETABS internal design combo codes (DStlSXX) are treated as 'gravity'
    because the underlying load pattern is opaque.
    """
    if not combo_name or combo_name in ('nan', 'None', ''):
        return 'gravity'
    if re.match(r'^DStl', combo_name, re.IGNORECASE):
        return 'gravity'
    return 'lateral' if LATERAL_RE.search(combo_name) else 'gravity'


# ---------------------------------------------------------------------------
# % change computation
# ---------------------------------------------------------------------------

def _pct_change(exist_val: float, new_val: float):
    """
    Compute percent change in magnitude: (|new| - |exist|) / |exist| * 100.

    Returns (pct, sign_reversal) where pct is a float, 'INF', or 0.0.
    sign_reversal is True when both values are non-zero and have opposite signs.
    """
    near_zero = 1e-4
    e_abs = abs(exist_val)
    n_abs = abs(new_val)

    sign_rev = (e_abs >= near_zero and n_abs >= near_zero and exist_val * new_val < 0)

    if e_abs < near_zero and n_abs < near_zero:
        return 0.0, False
    if e_abs < near_zero:
        return 'INF', sign_rev

    pct = (n_abs - e_abs) / e_abs * 100.0
    return round(pct, 1), sign_rev


def _fmt(val, decimals=3):
    """Format a numeric value for display, or pass through non-numeric strings."""
    if isinstance(val, float) and not (val != val):  # not NaN
        return round(val, decimals)
    return val


def _gov_combo(series, forces: list) -> str:
    """Return the OutputCase that drove the max absolute value among the given governing forces."""
    if series is None:
        return 'N/A'
    best_val, best_combo = 0.0, ''
    for c in forces:
        try:
            v = abs(float(series[c]))
            if v > best_val:
                best_val = v
                best_combo = str(series.get(f'{c}_Combo', '') or '')
        except (TypeError, ValueError, KeyError):
            pass
    return best_combo or 'N/A'


# ---------------------------------------------------------------------------
# Fail reason helper
# ---------------------------------------------------------------------------

def _compute_fail_reason(row: dict, threshold: float, load_type: str,
                          any_sign_rev: bool, is_fail: bool, gov_forces: list) -> str:
    """
    Build a human-readable string for the primary cause of a FLAG/WARN.

    Only checks gov_forces — non-governing forces do not appear in the reason.
    Returns '' for passing or non-comparable rows.
    """
    if not is_fail:
        return ''

    # Priority 1: governing force jumped from ~zero (INF change)
    for c in gov_forces:
        if row.get(f'{c}_Pct') == 'INF':
            return f'{c} INF ({load_type})'

    # Priority 2: largest % overage among governing forces
    worst_label = None
    worst_pct = 0.0
    for c in gov_forces:
        pct = row.get(f'{c}_Pct')
        if isinstance(pct, float) and pct > threshold and pct > worst_pct:
            worst_pct = pct
            worst_label = c

    if worst_label is not None:
        sign_tag = ' [sign rev]' if any_sign_rev else ''
        return f'{worst_label} +{worst_pct:.1f}% > {threshold:.0f}% {load_type}{sign_tag}'

    return 'See detail'


# ---------------------------------------------------------------------------
# Internal comparison (no file I/O — works from pre-parsed dicts)
# ---------------------------------------------------------------------------

def _run_comparison_internal(
    parsed_exist: dict,
    parsed_new: dict,
    member_type_filter: str,
    gravity_threshold: float,
    lateral_threshold: float,
    gravity_warn_threshold: float,
    lateral_warn_threshold: float,
) -> list:
    """
    Core comparison logic. Operates on pre-parsed DataFrames.
    Returns the full unfiltered result list.

    Flagging uses only GOVERNING_FORCES per member type; PMM D/C is computed
    and stored for display but does not drive FLAG/WARN status.
    Load type (gravity/lateral) is derived from the per-force governing combo,
    with fallback to PMMCombo from the design summary.
    """
    types_to_process = (
        ['Columns', 'Beams', 'Braces']
        if member_type_filter == 'All'
        else [member_type_filter]
    )

    sum_exist = parsed_exist['summary']
    sum_new   = parsed_new['summary']

    all_results = []

    for mtype in types_to_process:
        design_type = DESIGN_TYPE[mtype]

        ds_exist = (
            sum_exist[sum_exist['MemberType'] == design_type].copy()
            if 'MemberType' in sum_exist.columns else pd.DataFrame()
        )
        ds_new = (
            sum_new[sum_new['MemberType'] == design_type].copy()
            if 'MemberType' in sum_new.columns else pd.DataFrame()
        )

        forces_exist = parsed_exist['forces'][mtype]
        forces_new   = parsed_new['forces'][mtype]

        def _force_member_set(df):
            if df.empty or 'Story' not in df.columns:
                return set()
            return set(zip(df['Story'], df['Label']))

        exist_members = _force_member_set(forces_exist)
        new_members   = _force_member_set(forces_new)
        all_members   = exist_members | new_members

        def _build_lookup(ds):
            out = {}
            if ds.empty:
                return out
            for _, r in ds.iterrows():
                out[(r['Story'], r['Label'])] = r
            return out

        exist_lookup = _build_lookup(ds_exist)
        new_lookup   = _build_lookup(ds_new)

        def _build_force_lookup(df):
            out = {}
            if df.empty:
                return out
            for _, r in df.iterrows():
                out[(r['Story'], r['Label'])] = r
            return out

        fe_lookup = _build_force_lookup(forces_exist)
        fn_lookup = _build_force_lookup(forces_new)

        gov_forces = GOVERNING_FORCES[mtype]

        for story, label in sorted(all_members, key=lambda x: (str(x[0]), str(x[1]))):
            in_exist = (story, label) in exist_members
            in_new   = (story, label) in new_members

            es = exist_lookup.get((story, label))   # summary row — may be None
            ns = new_lookup.get((story, label))     # summary row — may be None
            ef = fe_lookup.get((story, label))
            nf = fn_lookup.get((story, label))

            row = {
                'Story':      story,
                'Label':      label,
                'MemberType': design_type,
            }

            # ---- REMOVED ----
            if not in_new:
                row.update({
                    'DesignSection_Exist': es.get('DesignSection', '') if es is not None else '',
                    'DesignSection_New':   'N/A',
                    'GovCombo_Exist': _gov_combo(ef, gov_forces) if ef is not None else (es.get('PMMCombo', '') if es is not None else ''),
                    'GovCombo_New':   'N/A',
                    'LoadType': 'N/A',
                    'PMM_Exist': _fmt(es.get('PMMRatio')) if es is not None else 'N/A',
                    'PMM_New': 'N/A', 'PMM_Pct': 'N/A',
                    'VMaj_Exist': _fmt(es.get('VMajRatio')) if es is not None else 'N/A',
                    'VMaj_New': 'N/A', 'VMaj_Pct': 'N/A',
                    'SignReversal': '',
                    'FailReason': '',
                    'Pass': 'REMOVED',
                })
                for c in FORCE_COMPONENTS:
                    row[f'{c}_Exist'] = _fmt(float(ef[c]), 2) if ef is not None else 'N/A'
                    row[f'{c}_New']   = 'N/A'
                    row[f'{c}_Pct']   = 'N/A'
                all_results.append(row)
                continue

            # ---- ADDED ----
            if not in_exist:
                n_gov_c = _gov_combo(nf, gov_forces) if nf is not None else 'N/A'
                n_load_type = classify_combo(n_gov_c if n_gov_c != 'N/A' else (ns.get('PMMCombo', '') if ns is not None else ''))
                row.update({
                    'DesignSection_Exist': 'N/A',
                    'DesignSection_New':   ns.get('DesignSection', '') if ns is not None else '',
                    'GovCombo_Exist': 'N/A',
                    'GovCombo_New': n_gov_c,
                    'LoadType': n_load_type,
                    'PMM_Exist': 'N/A',
                    'PMM_New': _fmt(ns.get('PMMRatio')) if ns is not None else 'N/A',
                    'PMM_Pct': 'N/A',
                    'VMaj_Exist': 'N/A',
                    'VMaj_New': _fmt(ns.get('VMajRatio')) if ns is not None else 'N/A',
                    'VMaj_Pct': 'N/A',
                    'SignReversal': '',
                    'FailReason': '',
                    'Pass': 'ADDED',
                })
                for c in FORCE_COMPONENTS:
                    row[f'{c}_Exist'] = 'N/A'
                    row[f'{c}_New']   = _fmt(float(nf[c]), 2) if nf is not None else 'N/A'
                    row[f'{c}_Pct']   = 'N/A'
                all_results.append(row)
                continue

            # ---- MATCHED ----
            e_sec  = es.get('DesignSection', '') if es is not None else ''
            n_sec  = ns.get('DesignSection', '') if ns is not None else ''
            e_pmm  = float(es['PMMRatio'])  if es is not None and pd.notna(es.get('PMMRatio'))  else None
            n_pmm  = float(ns['PMMRatio'])  if ns is not None and pd.notna(ns.get('PMMRatio'))  else None
            e_vmaj = float(es['VMajRatio']) if es is not None and pd.notna(es.get('VMajRatio')) else None
            n_vmaj = float(ns['VMajRatio']) if ns is not None and pd.notna(ns.get('VMajRatio')) else None

            # Governing combos from per-force data; fall back to PMMCombo if unavailable
            e_gov_combo = _gov_combo(ef, gov_forces) if ef is not None else (es.get('PMMCombo', '') if es is not None else 'N/A')
            n_gov_combo = _gov_combo(nf, gov_forces) if nf is not None else (ns.get('PMMCombo', '') if ns is not None else 'N/A')
            # Overall load type from governing combo — used as display fallback only
            overall_load_type = classify_combo(n_gov_combo if n_gov_combo != 'N/A' else (ns.get('PMMCombo', '') if ns is not None else ''))

            row['DesignSection_Exist'] = e_sec
            row['DesignSection_New']   = n_sec
            row['GovCombo_Exist']      = e_gov_combo
            row['GovCombo_New']        = n_gov_combo

            # D/C ratios — computed for display; do not drive FLAG/WARN
            pmm_pct  = _pct_change(e_pmm,  n_pmm)[0]  if e_pmm  is not None and n_pmm  is not None else 'N/A'
            vmaj_pct = _pct_change(e_vmaj, n_vmaj)[0] if e_vmaj is not None and n_vmaj is not None else 'N/A'

            row['PMM_Exist']  = _fmt(e_pmm)  if e_pmm  is not None else 'N/A'
            row['PMM_New']    = _fmt(n_pmm)  if n_pmm  is not None else 'N/A'
            row['PMM_Pct']    = pmm_pct
            row['VMaj_Exist'] = _fmt(e_vmaj) if e_vmaj is not None else 'N/A'
            row['VMaj_New']   = _fmt(n_vmaj) if n_vmaj is not None else 'N/A'
            row['VMaj_Pct']   = vmaj_pct

            any_sign_rev = False
            is_fail = False
            is_warn = False
            # worst_fail / worst_warn: keyed by load type → (force, numeric_pct, raw_pct_val)
            # Tracks the worst flagging force independently per load type so both
            # gravity and lateral overages can surface in the FailReason.
            worst_fail: dict = {}
            worst_warn: dict = {}

            for c in FORCE_COMPONENTS:
                e_val = _fmt(float(ef[c]), 2) if ef is not None else 'No Data'
                n_val = _fmt(float(nf[c]), 2) if nf is not None else 'No Data'

                row[f'{c}_Exist'] = e_val
                row[f'{c}_New']   = n_val

                if isinstance(e_val, (int, float)) and isinstance(n_val, (int, float)):
                    pct, sign_rev = _pct_change(float(e_val), float(n_val))
                    row[f'{c}_Pct'] = pct
                    if sign_rev:
                        any_sign_rev = True
                    # Each governing force evaluated against the threshold for its own combo's load type
                    if c in gov_forces:
                        c_combo = str(nf.get(f'{c}_Combo', '') or '') if nf is not None else ''
                        c_load_type = classify_combo(c_combo)
                        row[f'{c}_LoadType'] = c_load_type
                        c_fail_thresh = lateral_threshold      if c_load_type == 'lateral' else gravity_threshold
                        c_warn_thresh = lateral_warn_threshold if c_load_type == 'lateral' else gravity_warn_threshold
                        numeric_pct = 999.0 if pct == 'INF' else (pct if isinstance(pct, float) else None)
                        if numeric_pct is not None:
                            if numeric_pct > c_fail_thresh:
                                is_fail = True
                                if numeric_pct > worst_fail.get(c_load_type, (None, 0.0, None))[1]:
                                    worst_fail[c_load_type] = (c, numeric_pct, pct)
                            elif numeric_pct > c_warn_thresh:
                                is_warn = True
                                if numeric_pct > worst_warn.get(c_load_type, (None, 0.0, None))[1]:
                                    worst_warn[c_load_type] = (c, numeric_pct, pct)
                else:
                    row[f'{c}_Pct'] = 'N/A'

            row['SignReversal'] = 'YES' if any_sign_rev else ''

            def _reason_parts(worst_dict: dict, thresh_fn) -> list:
                """Build one reason string per load type, sorted lateral-first."""
                parts = []
                for lt in sorted(worst_dict, key=lambda x: (x != 'lateral')):
                    c, _, pv = worst_dict[lt]
                    ft = thresh_fn(lt)
                    parts.append(
                        f'{c} INF ({lt})' if pv == 'INF'
                        else f'{c} +{pv:.1f}% > {ft:.0f}% {lt}'
                    )
                return parts

            sign_tag = ' [sign rev]' if any_sign_rev else ''

            # LoadType on the row: load type of the highest-overage flagging force
            # (used for gravity/lateral counts in the summary; 'mixed' when both flag)
            def _primary_load_type(worst_dict: dict) -> str:
                if not worst_dict:
                    return overall_load_type
                if len(worst_dict) == 1:
                    return next(iter(worst_dict))
                return max(worst_dict, key=lambda lt: worst_dict[lt][1])

            if is_fail:
                parts = _reason_parts(worst_fail, lambda lt: lateral_threshold if lt == 'lateral' else gravity_threshold)
                row['FailReason'] = ' | '.join(parts) + sign_tag
                row['LoadType']   = _primary_load_type(worst_fail)
                row['Pass']       = 'FLAG'
            elif is_warn:
                parts = _reason_parts(worst_warn, lambda lt: lateral_warn_threshold if lt == 'lateral' else gravity_warn_threshold)
                row['FailReason'] = ' | '.join(parts) + sign_tag
                row['LoadType']   = _primary_load_type(worst_warn)
                row['Pass']       = 'WARN'
            else:
                row['FailReason'] = ''
                row['LoadType']   = overall_load_type
                row['Pass']       = 'PASS'

            all_results.append(row)

    return all_results


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def run_comparison(
    existing_file,
    modified_file,
    member_type_filter: str,
    gravity_threshold: float,
    lateral_threshold: float,
    gravity_warn_threshold: float = 2.5,
    lateral_warn_threshold: float = 5.0,
    show_failures_only: bool = False,
    _parsed_new_override: dict = None,
    _cache_key_extra: tuple = (),
) -> list:
    """
    Compare ETABS design output between two models.

    Returns a list of row dicts — one per matched (or unmatched) member.
    Results are cached by file hash + comparison parameters; only file parsing
    (on first upload) is slow.

    Keys per row: Story, Label, MemberType, DesignSection_Exist/New,
    GovCombo_Exist/New, LoadType, P/V2/V3/T/M2/M3_Exist/New/Pct,
    PMM_Exist/New/Pct, VMaj_Exist/New/Pct, SignReversal, FailReason, Pass

    _parsed_new_override: pre-processed parsed_dict for the new model (e.g. with
        labels remapped for coordinate matching). When None the default parse is used.
    _cache_key_extra: additional tuple components appended to the cache key so
        different matching modes produce separate cache entries.
    """
    exist_hash, parsed_exist = _get_parsed_file(existing_file)
    new_hash,   parsed_new   = _get_parsed_file(modified_file)
    if _parsed_new_override is not None:
        parsed_new = _parsed_new_override
    cache_key  = (exist_hash, new_hash, member_type_filter,
                  gravity_threshold, lateral_threshold,
                  gravity_warn_threshold, lateral_warn_threshold) + _cache_key_extra

    all_results = get_cached('etabs_cmpv4', cache_key)
    if all_results is None:
        all_results = _run_comparison_internal(
            parsed_exist, parsed_new,
            member_type_filter, gravity_threshold, lateral_threshold,
            gravity_warn_threshold, lateral_warn_threshold,
        )
        set_cached('etabs_cmpv4', cache_key, all_results)

    if show_failures_only:
        return [r for r in all_results if r.get('Pass') in ('FLAG', 'WARN')]
    return list(all_results)


def build_summary(results: list) -> list:
    """
    Group results by MemberType and Story; count PASS/FAIL/ADDED/REMOVED.
    Returns list of row dicts: Story, MemberType, Total, PASS, FAIL, ADDED, REMOVED.
    """
    counts = defaultdict(lambda: {'Total': 0, 'PASS': 0, 'WARN': 0, 'FLAG': 0, 'ADDED': 0, 'REMOVED': 0})

    for r in results:
        key = (r.get('Story', ''), r.get('MemberType', ''))
        counts[key]['Total'] += 1
        status = r.get('Pass', '')
        if status in ('PASS', 'WARN', 'FLAG', 'ADDED', 'REMOVED'):
            counts[key][status] += 1

    summary = []
    for (story, mtype), c in sorted(counts.items(), key=lambda x: (str(x[0][0]), str(x[0][1]))):
        summary.append({
            'Story':      story,
            'MemberType': mtype,
            'Total':      c['Total'],
            'PASS':       c['PASS'],
            'WARN':       c['WARN'],
            'FLAG':       c['FLAG'],
            'ADDED':      c['ADDED'],
            'REMOVED':    c['REMOVED'],
        })
    return summary


def results_to_csv(results: list) -> str:
    """Convert results list-of-dicts to a CSV string."""
    if not results:
        return 'No results\n'
    return pd.DataFrame(results).to_csv(index=False)
